# # 列表去重方法
# book_name_list = ["吴起", "西峰", "安塞", "吴起", "吴起"]
# print(book_name_list)
# new_book_name_list = set(book_name_list)
# print(new_book_name_list)
# # 通过子站地址查询安装位置
# a_dict = {'333': "悦乐111线13杆", '335': "悦乐112线1杆"}
# b = input("请输入子站地址")
# print(f"子站地址{b},"+"安装位置为"+a_dict.get(b, "当前子站地址不存在"))
import openpyxl as vb
import openpyxl as vb
workbook_name = input("请输入用户单位名称，；例如西峰电力大队，输入西峰 ：")
fls_sheet_name= "通讯终端"
efs_sheet_name = "信号源"
fls_number = int(input("请输入通讯终端数量：")) + 2
efs_number = int(input("请输入信号源的数量：")) + 2
print("输入完成")
# 工作簿路径
path = f"{workbook_name}电力大队故障定位设备台账信息.xlsx"
print("路径定义成功")
# 将工作簿路径赋予wb
wb = vb.load_workbook(path)
gprs_fls_name_list = []
gprs_fls_number_list = []
gprs_fls_modemid_list = []
gprs_fls_port_list =[]
gprs_efs_name_list =[]
gprs_efs_number_list =[]
gprs_efs_modemid_list = []
gprs_efs_port_list =[]
i = 2
j = 2
print("名称及地址列表定义成功")
while i < fls_number:
    # 定义通讯终端名字
    fls_ws = wb[f'{fls_sheet_name}']
    # c = fls_ws.cell(row=i, column=10).value
    gprs_fls_name_list.append(fls_ws.cell(row=i, column=10).value)
    # d = fls_ws.cell(row=i, column=11).value
    gprs_fls_number_list.append(fls_ws.cell(row=i, column=11).value)
    gprs_fls_modemid_list.append(fls_ws.cell(row=i, column=12).value)
    gprs_fls_port_list.append(fls_ws.cell(row=i, column=13).value)
    i += 1
while j < efs_number:
    efs_ws = wb[f'{efs_sheet_name}']
    gprs_efs_name_list.append(efs_ws.cell(row=j, column=3).value)
    gprs_efs_number_list.append(efs_ws.cell(row=j, column=4).value)
    gprs_efs_modemid_list.append(efs_ws.cell(row=j, column=5).value)
    gprs_efs_port_list.append(efs_ws.cell(row=j, column=6).value)
    j += 1
print(gprs_fls_name_list)
print(gprs_fls_number_list)
print("通讯终端名称及地址列表添加成功")
print(gprs_efs_name_list)
print(gprs_efs_number_list)
print("信号源名称及地址列表添加成功")
print(gprs_fls_modemid_list)
print(gprs_fls_port_list)
print("通讯终端MODEMID及通道和列表添加成功")
print(gprs_efs_modemid_list)
print(gprs_efs_port_list)
print("信号源MODEMID及通道号列表添加成功")
a = 2
c = 0
sms_center = '''[citGPRS FAC 1]
NAME=短信中心1
FACADDR=1
MODEMID=00000001
SVR=短信中心
COM=COM1
TRAN=COM1
SUBFACNO=1
'''
with open(r'gprs.ini', 'w') as f:
    f.write(sms_center)
    for i in gprs_fls_port_list:
        if i == 5002:
            f.write(f'''\n[citGPRS FAC {a}]
NAME={gprs_fls_name_list[c]}
FACADDR={gprs_fls_number_list[c]}
MODEMID={gprs_fls_modemid_list[c]}
SVR=通讯终端GPRS2
COM=COM4
TRAN=COM1
SUBFACNO={gprs_fls_number_list[c]}
''')
        else:
            f.write(f'''\n[citGPRS FAC {a}]
NAME={gprs_fls_name_list[c]}
FACADDR={gprs_fls_number_list[c]}
MODEMID=00{gprs_fls_modemid_list[c]}
SVR=通讯终端GPRS
COM=COM2
TRAN=COM1
SUBFACNO={gprs_fls_number_list[c]}
''')
        a += 1

        c += 1
    c = 0
    for j in gprs_efs_number_list:
        f.write(f'''\n[citGPRS FAC {a}]
NAME={gprs_efs_name_list[c]}
FACADDR={j}
MODEMID=00{j}
SVR=信号源GPRS
COM=COM3
SUBFACNO={j}
''')
        a += 1

        c += 1
