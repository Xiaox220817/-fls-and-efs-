from openpyxl import Workbook
from openpyxl.formula.translate import Translator

book_name_list = []  # 创建一个空列表用来存放用户单位的名称


def new_name_list():
    # 这个函数可以让操作者输入需要创建台账的单位的名称
    global book_name_list  # 声明book_name_list是一个全局变量
    book_name_list = []  # 再次清空book_name_list列表

    while True:
        # 输入用户单位名称
        book_name = input("请输入一个用户单位名称，输入 'finish' 结束,请开始输入:")
        # 通过if函数条件判断来结束循环
        if book_name == "finish":  # 输入finish结束输入
            break  # 结束循环
        else:
            # 将使用者输入的名称先转为字符串格式，然后存储到book_name_list列表中
            book_name_list.append(book_name)
    print(f"您输入的单位名称有：{book_name_list}")


# 调用new_name_list()函数，将输入的名称存放至book_name_list中
new_name_list()


def new_workbook():
    # 这个函数用来新建excel表格
    def new_fls_sheet_formula():
        # 这个函数用来设置故障定位终端表内的表头及格式
        # 输入通讯终端的数量,调用创建excel文件时for循环的变量i
        fls_number = int(input(f"请输入{i}的通讯终端数量，数量不可为小数："))
        # 在工作表内写入列表名
        fls_sheet_name = ['序号', '变电站', '线路', '杆号', '是否电源侧', '是否电源侧结果', '是否分支',
                          '是否分支结果',
                          '名称（concatenate格式）', '安装人员输入杆号', '子站地址', '通道号', '通讯类型',
                          '出厂日期', '终端版本号',
                          'RF模块版本', '指示器型号', '指示器参数', '通讯模块品牌', '通讯模块版本号',
                          '太阳能板电压及功率', '电池电压及容量',
                          '电容电压及容量', 'SIM运营商', 'SIM卡IP', 'ICCID', 'APN', '登录帧', ' 备注']
        ws.append(fls_sheet_name)
        ws['A2'] = '=ROW()-1'  # 序号公式
        ws['B2'] = '=IFERROR(LEFT(J2,FIND("变",J2,1)),"未安装")'  # 获取变电站名公式，未获取到自动填入未安装
        ws['C2'] = '=IFERROR(MID(J2,FIND("变",J2)+1,FIND("线",J2)-FIND("变",J2)),"未安装")'  # 获取线路名称公式，未获取到自动填入未安装
        ws[
            'D2'] = '=IFERROR(MID(J2,FIND("线",J2)+1,FIND("杆",J2)-(FIND("线",J2)+1)),"未安装")'  # 获取杆号公式，未获取到自动填入未安装
        ws['E2'] = '=IFERROR(IF(FIND("电源侧",J2,1),"是"),"否")'  # 判断电源侧侧安装公式
        ws['F2'] = '=IF(E2="是","电源侧","")'  # 是否电源侧安装结果
        ws['G2'] = '=IFERROR(IF(FIND("分支",J2,1),"是"),"否")'  # 判断是否分支侧安装公式
        ws['H2'] = '=IF(G2="是","分支","")'  # 是否分支侧安装结果
        ws['I2'] = '=IFERROR(CONCATENATE(B2,C2,D2,F2,H2),"未安装")'  # 根据前边的结果生成杆号
        # 在A3:A50内填充上边所写的公式，由于是闭区间，所以需要输入的设备数量加1
        for cell in ws[f"A3:A{fls_number + 1}"]:
            cell[0].value = Translator(formula='=ROW()-1', origin='A2').translate_formula(cell[0].coordinate)
        for cell in ws[f"B3:B{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IFERROR(LEFT(J2,FIND("变",J2,1)),"未安装")',
                                       origin='B2').translate_formula(cell[0].coordinate)
        for cell in ws[f"C3:C{fls_number + 1}"]:
            cell[0].value = Translator(
                formula='=IFERROR(MID(J2,FIND("变",J2)+1,FIND("线",J2)-FIND("变",J2)),"未安装")',
                origin='C2').translate_formula(cell[0].coordinate)
        for cell in ws[f"D3:D{fls_number + 1}"]:
            cell[0].value = Translator(
                formula='=IFERROR(MID(J2,FIND("线",J2)+1,FIND("杆",J2)-(FIND("线",J2)+1)),"未安装")',
                origin='D2').translate_formula(cell[0].coordinate)
        for cell in ws[f"E3:E{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IFERROR(IF(FIND("电源侧",J2,1),"是"),"否")',
                                       origin='E2').translate_formula(cell[0].coordinate)
        for cell in ws[f"F3:F{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IF(E2="是","电源侧","")', origin='F2').translate_formula(
                cell[0].coordinate)
        for cell in ws[f"G3:G{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IFERROR(IF(FIND("分支",J2,1),"是"),"否")',
                                       origin='G2').translate_formula(cell[0].coordinate)
        for cell in ws[f"H3:H{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IF(G2="是","分支","")', origin='H2').translate_formula(
                cell[0].coordinate)
        for cell in ws[f"I3:I{fls_number + 1}"]:
            cell[0].value = Translator(formula='=IFERROR(CONCATENATE(B2,C2,D2,F2,H2),"未安装")',
                                       origin='I2').translate_formula(cell[0].coordinate)

    def new_efs_sheet_formula():
        # 这个函数用来设置信号源表格
        # 输入通讯终端的数量，调用创建excel文件时for循环的变量i
        efs_number = int(input(f"请输入{i}的信号源数量，数量不可为小数："))
        # 设置表头
        efs_sheet = ['序号', '变电所', '名称', '子站地址', '通讯方式', '出厂日期', '产品型号', '版本号',
                     '液晶版本号', '故障延迟时间', 'CT变比', '接地脉冲间隔1', '接地脉冲间隔2', 'SIM卡IP']
        ws1.append(efs_sheet)
        ws1['A2'] = '=ROW()-1'
        ws1['B2'] = '=IFERROR(LEFT(C2,FIND("变",C2,1)),"未安装")'
        for cell1 in ws1[f"A3:A{efs_number + 1}"]:
            cell1[0].value = Translator(formula='=ROW()-1', origin='A2').translate_formula(cell1[0].coordinate)
        for cell1 in ws1[f"B3:B{efs_number + 1}"]:
            cell1[0].value = Translator(formula='=IFERROR(LEFT(C2,FIND("变",C2,1)),"未安装")',
                                        origin='B2').translate_formula(cell1[0].coordinate)

    successfully_created_list = []  # 创建一个列表，用来存放台账创建成功的单位的名称

    # 遍历booK_name_list列表，使用列表的内容为名称创建excel
    for i in book_name_list:
        wb = Workbook()  # 创建一个excel表格
        ws = wb.active  # ws为name.excel的默认表
        ws.title = "FLS_sheet"  # 将默认表的表名改为FLS_sheet
        # 在工作簿中创建一个EFS_sheet的表，位置为2
        ws1 = wb.create_sheet("EFS_sheet", 1)
        # 在工作簿中创建一个Device_statistics_sheet的表，位置为3
        ws2 = wb.create_sheet("Device_statistics_sheet", 2)
        # 调用函数new_equipment_sheet_formula()，在equipment_sheet中添加表头和公式
        new_fls_sheet_formula()
        # 调用函数new_signal_source_sheet_formula()，在esignal_source_sheet中添加表头和公式
        new_efs_sheet_formula()
        # 保存工作簿，名称为book_name_list中下标为0的值+电力大队设备台账.xlsx
        wb.save(f"{i}故障定位系统设备台账.xlsx")  # 存储excel表格
        # 告知用户保存工作簿成功
        successfully_created_list.append(i)  # 将本次使用的单位名称存入列表
        # 遍历工作簿内得表名
        print(f"{i}故障定位系统设备台账内的工作表有：{wb.sheetnames}")
    print(f"已创建{len(successfully_created_list)}个设备台账")  # 告知操作者本次创建了几个台账
    print(f"本次成功创建了{successfully_created_list}的设备台账")  # 告知操作者本次创建了哪些台账


# 调用workbook函数，创建excel表格
new_workbook()
